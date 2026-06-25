#!/usr/bin/env python
"""Per-record change report between two Iceberg snapshots of one lake table.

Compares the table's **latest** snapshot against the **last snapshot of a given
day** (default: yesterday) and reports, per record, which business columns
changed. Records are identified by ``(branch_id, <unique_key>)``; a record is
"updated" when it exists in both snapshots but at least one business column's
*canonical* value differs. ETL-injected columns (``insert_at``,
``recorded_updated_at``, ``_dlt*``) are excluded so a plain reload that only
re-stamps load times is not reported as a change.

Value comparison reuses the pipeline's own canonicalizer (``etl.dq_check``) so a
timestamp/decimal/number that merely *reprints* differently is treated as equal
-- only real content drift is flagged.

Outputs (under ``--out-dir``, default ``exports/``):
  * ``<table>_changes_<asof>_vs_<latest>.csv``   -- tidy: one row per changed field
  * ``<table>_changes_<asof>_vs_<latest>.xlsx``  -- Summary + Updated Records
    (changed cells highlighted, shown as ``old -> new``)

Example
-------
    python snapshot_diff.py                         # delivery_charge, yesterday vs latest
    python snapshot_diff.py --table delivery_charge --as-of 2026-06-23
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from pathlib import Path
from typing import Iterator, Optional

import pyarrow as pa
import pyarrow.compute as pc

from etl.dq_check import _canon_array, _fingerprint, _norm, _resolve_actual

# Identity / injected columns (normalized lake names) never treated as "business".
INJECTED = {"insert_at", "recorded_updated_at"}
LAKE_ROOT_DEFAULT = Path("iceberg_output") / "oasis"
KEY_SEP = "\x1f"


# --------------------------------------------------------------------------- #
# Snapshot selection
# --------------------------------------------------------------------------- #
def _open_table(root: Path, table: str):
    meta_dir = root / table / "metadata"
    metas = list(meta_dir.glob("*.metadata.json"))
    if not metas:
        raise SystemExit(f"No Iceberg metadata for {table!r} under {meta_dir}")

    def ver(p: Path) -> int:
        m = re.match(r"(\d+)-", p.name)
        return int(m.group(1)) if m else -1

    meta = max(metas, key=lambda p: (ver(p), p.stat().st_mtime))
    from pyiceberg.table import StaticTable

    return StaticTable.from_metadata("file://" + str(meta.resolve()).replace("\\", "/"))


def _snapshots_sorted(table) -> list:
    """All snapshots, oldest-first by commit time."""
    return sorted(table.metadata.snapshots, key=lambda s: s.timestamp_ms)


def _pick_snapshots(table, as_of: dt.date) -> tuple[object, object]:
    """(as_of_last, latest): last snapshot committed on ``as_of`` local date, and
    the most recent snapshot overall."""
    snaps = _snapshots_sorted(table)
    if not snaps:
        raise SystemExit("Table has no snapshots.")
    latest = snaps[-1]
    on_day = [s for s in snaps
              if dt.datetime.fromtimestamp(s.timestamp_ms / 1000).date() == as_of]
    if not on_day:
        days = sorted({dt.datetime.fromtimestamp(s.timestamp_ms / 1000).date()
                       for s in snaps})
        raise SystemExit(
            f"No snapshot committed on {as_of}. Snapshots exist for: "
            + ", ".join(str(d) for d in days))
    as_of_last = on_day[-1]
    if as_of_last.snapshot_id == latest.snapshot_id:
        raise SystemExit(
            f"The last snapshot of {as_of} *is* the latest snapshot "
            f"({latest.snapshot_id}); nothing to compare.")
    return as_of_last, latest


def _local(ts_ms: int) -> str:
    return dt.datetime.fromtimestamp(ts_ms / 1000).strftime("%Y-%m-%d %H:%M:%S")


# --------------------------------------------------------------------------- #
# Column model
# --------------------------------------------------------------------------- #
def _column_model(table, unique_key_norm: str) -> tuple[str, list[str]]:
    """Return (branch_col, business_cols) using actual lake field names."""
    names = [f.name for f in table.schema().fields]
    branch_col = _resolve_actual(names, "branch_id") or "branch_id"
    key_col = _resolve_actual(names, unique_key_norm)
    if key_col is None:
        raise SystemExit(f"Unique key {unique_key_norm!r} not found in {table}")
    identity = {branch_col, key_col}
    business = [n for n in names
                if n not in identity
                and _norm(n) not in INJECTED
                and not n.startswith("_dlt")]
    return branch_col, key_col, business


# --------------------------------------------------------------------------- #
# Scanning a specific snapshot
# --------------------------------------------------------------------------- #
def _scan(table, snapshot_id: int, columns: list[str]) -> Iterator[pa.Table]:
    scan = table.scan(snapshot_id=snapshot_id, selected_fields=tuple(columns))
    for rb in scan.to_arrow_batch_reader():
        if rb.num_rows:
            yield pa.Table.from_batches([rb])


def _key_hash_table(table, snapshot_id: int, branch_col: str, key_col: str,
                    business: list[str]) -> tuple[pa.Table, int, int]:
    """Stream a snapshot into a (key, hash) Arrow table. Returns (kh, rows, dup_keys)."""
    from hashlib import blake2b

    cols = [branch_col, key_col] + business
    keys_parts, hash_parts, rows = [], [], 0
    for batch in _scan(table, snapshot_id, cols):
        keys = _fingerprint(batch, [branch_col, key_col])
        payload = _fingerprint(batch, business)
        hashes = pa.array(
            [blake2b(p.encode("utf-8"), digest_size=16).hexdigest()
             for p in payload.to_pylist()], pa.string())
        keys_parts.append(keys)
        hash_parts.append(hashes)
        rows += batch.num_rows
    if not keys_parts:
        empty = pa.table({"k": pa.array([], pa.string()), "h": pa.array([], pa.string())})
        return empty, 0, 0
    kh = pa.table({"k": pa.chunked_array(keys_parts), "h": pa.chunked_array(hash_parts)})
    distinct = pc.count_distinct(kh.column("k")).as_py()
    return kh, rows, rows - distinct


def _diff_keys(old_kh: pa.Table, new_kh: pa.Table) -> tuple[pa.Array, int, int]:
    """Full-outer-join two (k,h) tables; return (updated_keys, n_inserted, n_deleted)."""
    o = old_kh.rename_columns(["k", "ho"])
    n = new_kh.rename_columns(["k", "hn"])
    joined = o.join(n, keys="k", join_type="full outer")
    ho, hn = joined.column("ho"), joined.column("hn")
    o_null, n_null = pc.is_null(ho), pc.is_null(hn)
    both = pc.and_(pc.invert(o_null), pc.invert(n_null))
    updated_mask = pc.and_(both, pc.invert(pc.equal(ho, hn)))
    updated_keys = joined.filter(updated_mask).column("k").combine_chunks()
    n_inserted = pc.sum(pc.cast(pc.and_(o_null, pc.invert(n_null)), pa.int64())).as_py() or 0
    n_deleted = pc.sum(pc.cast(pc.and_(pc.invert(o_null), n_null), pa.int64())).as_py() or 0
    return updated_keys, n_inserted, n_deleted


# --------------------------------------------------------------------------- #
# Materialize the changed rows from both snapshots (filtered to updated keys)
# --------------------------------------------------------------------------- #
def _collect_changed(table, snapshot_id: int, branch_col: str, key_col: str,
                     business: list[str], updated_keys: pa.Array) -> pa.Table:
    cols = [branch_col, key_col] + business
    keyset = pa.array(updated_keys, pa.string())
    out = []
    for batch in _scan(table, snapshot_id, cols):
        keys = _fingerprint(batch, [branch_col, key_col])
        mask = pc.is_in(keys, value_set=keyset)
        if pc.any(mask).as_py():
            sub = batch.filter(mask)
            sub = sub.append_column("__key", _fingerprint(sub, [branch_col, key_col]))
            out.append(sub)
    if not out:
        return pa.table({c: pa.array([], table.schema().as_arrow().field(c).type)
                         for c in cols} | {"__key": pa.array([], pa.string())})
    return pa.concat_tables(out)


def _align(tbl: pa.Table, ordered_keys: list[str]) -> pa.Table:
    pos = {k: i for i, k in enumerate(tbl.column("__key").to_pylist())}
    idx = pa.array([pos[k] for k in ordered_keys], pa.int64())
    return tbl.take(idx)


# --------------------------------------------------------------------------- #
# Report building
# --------------------------------------------------------------------------- #
def _fmt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def build_report(old_al: pa.Table, new_al: pa.Table, branch_col: str, key_col: str,
                 business: list[str], ordered_keys: list[str]):
    """Return (long_rows, changed_business_cols, per_col_counts)."""
    n = len(ordered_keys)
    branch_vals = new_al.column(branch_col).to_pylist()
    key_vals = new_al.column(key_col).to_pylist()

    canon = {c: (_canon_array(old_al.column(c)), _canon_array(new_al.column(c)))
             for c in business}
    raw_old = {c: old_al.column(c).to_pylist() for c in business}
    raw_new = {c: new_al.column(c).to_pylist() for c in business}
    changed_mask = {c: pc.not_equal(co, cn).to_pylist() for c, (co, cn) in canon.items()}

    per_col_counts: dict[str, int] = {}
    long_rows = []
    for i in range(n):
        for c in business:
            if changed_mask[c][i]:
                per_col_counts[c] = per_col_counts.get(c, 0) + 1
                long_rows.append({
                    "branch_id": branch_vals[i],
                    key_col: _fmt(key_vals[i]),
                    "column": c,
                    "old_value": _fmt(raw_old[c][i]),
                    "new_value": _fmt(raw_new[c][i]),
                })
    changed_cols = [c for c in business if c in per_col_counts]
    return long_rows, changed_cols, per_col_counts, raw_old, raw_new, changed_mask, branch_vals, key_vals


def write_excel(path: Path, *, table, branch_col, key_col, business, changed_cols,
                ordered_keys, raw_old, raw_new, changed_mask, branch_vals, key_vals,
                per_col_counts, n_updated, n_inserted, n_deleted, old_snap, new_snap,
                as_of, dup_old, dup_new):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    hi = PatternFill("solid", fgColor="FFF2A6")          # changed cell
    hdr = PatternFill("solid", fgColor="305496")
    hdr_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()
    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ["Table", table],
        ["Compared (baseline)", f"{as_of}  last snapshot {old_snap.snapshot_id}  @ {_local(old_snap.timestamp_ms)}"],
        ["Compared (latest)", f"snapshot {new_snap.snapshot_id}  @ {_local(new_snap.timestamp_ms)}"],
        [],
        ["Updated records (content changed)", n_updated],
        ["Inserted records (new in latest)", n_inserted],
        ["Deleted records (gone in latest)", n_deleted],
    ]
    if dup_old or dup_new:
        rows.append(["WARNING duplicate keys", f"baseline={dup_old}, latest={dup_new}"])
    rows += [[], ["Changed column", "# records affected"]]
    for c, cnt in sorted(per_col_counts.items(), key=lambda kv: -kv[1]):
        rows.append([c, cnt])
    for r in rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 70
    for cell in ("A1", "A2", "A3", "A5", "A6", "A7"):
        ws[cell].font = Font(bold=True)
    ws["A" + str(9 + (1 if (dup_old or dup_new) else 0))].font = Font(bold=True)

    # ---- Updated Records (wide; only columns that changed somewhere) ----
    ws2 = wb.create_sheet("Updated Records")
    header = ["branch_id", key_col, "# changes"] + changed_cols
    ws2.append(header)
    for j in range(len(header)):
        cell = ws2.cell(row=1, column=j + 1)
        cell.fill = hdr
        cell.font = hdr_font
        cell.alignment = wrap
    for i in range(len(ordered_keys)):
        nchg = sum(1 for c in changed_cols if changed_mask[c][i])
        row = [branch_vals[i], _fmt(key_vals[i]), nchg]
        for c in changed_cols:
            if changed_mask[c][i]:
                row.append(f"{_fmt(raw_old[c][i])}  →  {_fmt(raw_new[c][i])}")
            else:
                row.append(_fmt(raw_new[c][i]))
        ws2.append(row)
        excel_row = i + 2
        for k, c in enumerate(changed_cols):
            if changed_mask[c][i]:
                cell = ws2.cell(row=excel_row, column=4 + k)
                cell.fill = hi
    ws2.freeze_panes = "D2"
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 10
    for k in range(len(changed_cols)):
        ws2.column_dimensions[get_column_letter(4 + k)].width = 26
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(ordered_keys) + 1}"

    wb.save(path)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--table", default="delivery_charge", help="lake table name")
    p.add_argument("--unique-key", default="delivery_charge_id",
                   help="business unique key column (per branch)")
    p.add_argument("--as-of", help="baseline day YYYY-MM-DD (default: yesterday, local)")
    p.add_argument("--lake-root", default=str(LAKE_ROOT_DEFAULT),
                   help="path to <bucket>/<dataset>")
    p.add_argument("--out-dir", default="exports")
    args = p.parse_args(argv)

    as_of = (dt.date.fromisoformat(args.as_of) if args.as_of
             else dt.date.today() - dt.timedelta(days=1))
    root = Path(args.lake_root)
    table = _open_table(root, args.table)

    old_snap, new_snap = _pick_snapshots(table, as_of)
    print(f"Table         : {args.table}")
    print(f"Baseline (as-of {as_of}) : snapshot {old_snap.snapshot_id} @ {_local(old_snap.timestamp_ms)}")
    print(f"Latest                   : snapshot {new_snap.snapshot_id} @ {_local(new_snap.timestamp_ms)}")

    branch_col, key_col, business = _column_model(table, _norm(args.unique_key))
    print(f"Identity      : ({branch_col}, {key_col}) | {len(business)} business columns compared")

    print("Hashing baseline snapshot ...")
    old_kh, old_rows, dup_old = _key_hash_table(table, old_snap.snapshot_id, branch_col, key_col, business)
    print(f"  {old_rows:,} rows")
    print("Hashing latest snapshot ...")
    new_kh, new_rows, dup_new = _key_hash_table(table, new_snap.snapshot_id, branch_col, key_col, business)
    print(f"  {new_rows:,} rows")

    updated_keys, n_inserted, n_deleted = _diff_keys(old_kh, new_kh)
    n_updated = len(updated_keys)
    print(f"\nUpdated : {n_updated:,}   Inserted : {n_inserted:,}   Deleted : {n_deleted:,}")
    if dup_old or dup_new:
        print(f"WARNING: duplicate keys within a snapshot (baseline={dup_old}, latest={dup_new})")

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = f"{args.table}_changes_{as_of}_vs_{new_snap.snapshot_id}"
    csv_path = out_dir / f"{stem}.csv"
    xlsx_path = out_dir / f"{stem}.xlsx"

    if n_updated == 0:
        print("No updated records. Writing empty report.")
        csv_path.write_text("branch_id,%s,column,old_value,new_value\n" % key_col, encoding="utf-8")
        return 0

    print("Materializing changed rows from both snapshots ...")
    old_changed = _collect_changed(table, old_snap.snapshot_id, branch_col, key_col, business, updated_keys)
    new_changed = _collect_changed(table, new_snap.snapshot_id, branch_col, key_col, business, updated_keys)
    ordered_keys = sorted(set(updated_keys.to_pylist())
                          & set(old_changed.column("__key").to_pylist())
                          & set(new_changed.column("__key").to_pylist()))
    old_al = _align(old_changed, ordered_keys)
    new_al = _align(new_changed, ordered_keys)

    (long_rows, changed_cols, per_col_counts, raw_old, raw_new,
     changed_mask, branch_vals, key_vals) = build_report(
        old_al, new_al, branch_col, key_col, business, ordered_keys)

    # tidy CSV
    import csv as _csv
    with csv_path.open("w", newline="", encoding="utf-8-sig") as fh:
        w = _csv.DictWriter(fh, fieldnames=["branch_id", key_col, "column", "old_value", "new_value"])
        w.writeheader()
        w.writerows(long_rows)

    write_excel(xlsx_path, table=args.table, branch_col=branch_col, key_col=key_col,
                business=business, changed_cols=changed_cols, ordered_keys=ordered_keys,
                raw_old=raw_old, raw_new=raw_new, changed_mask=changed_mask,
                branch_vals=branch_vals, key_vals=key_vals, per_col_counts=per_col_counts,
                n_updated=n_updated, n_inserted=n_inserted, n_deleted=n_deleted,
                old_snap=old_snap, new_snap=new_snap, as_of=as_of,
                dup_old=dup_old, dup_new=dup_new)

    print(f"\nField-level changes : {len(long_rows):,} across {len(changed_cols)} column(s)")
    print("Top changed columns :")
    for c, cnt in sorted(per_col_counts.items(), key=lambda kv: -kv[1])[:10]:
        print(f"  {c:<28} {cnt:,}")
    print(f"\nWrote:\n  {csv_path}\n  {xlsx_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
